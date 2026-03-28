from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import os
import re
import traceback
from contextlib import asynccontextmanager
from dataclasses import dataclass, asdict
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import aiohttp
import aiosqlite
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import (
    CallbackQuery,
    ContentType,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
)
from aiogram.utils import executor
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv()

APP_NAME = "Xisobot Bot Pro"
DB_PATH = os.getenv("DB_PATH", "finance_bot.db")
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip()
GROQ_MODEL = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile").strip()
BOT_TIMEZONE = os.getenv("BOT_TIMEZONE", "Asia/Tashkent").strip()
DEFAULT_USD_RATE = Decimal(os.getenv("DEFAULT_USD_RATE", "12750").strip())
ADMIN_IDS = {
    int(x.strip())
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x.strip().isdigit()
}
RATE_API_URL = os.getenv("RATE_API_URL", "https://cbu.uz/uz/arkhiv-kursov-valyut/json/USD/")
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger(APP_NAME)

try:
    TZ = ZoneInfo(BOT_TIMEZONE)
except ZoneInfoNotFoundError:
    TZ = timezone(timedelta(hours=5))

bot = Bot(token=TELEGRAM_BOT_TOKEN, parse_mode="HTML")
dp = Dispatcher(bot, storage=MemoryStorage())

PENDING_BATCHES: Dict[int, Dict[str, Any]] = {}


# ------------------------------
# Formatting helpers
# ------------------------------
def now_tz() -> datetime:
    return datetime.now(TZ)


def dt_to_str(dt: datetime) -> str:
    return dt.astimezone(TZ).strftime("%d.%m.%Y %H:%M")


def today_str() -> str:
    return now_tz().strftime("%Y-%m-%d")


def decimal_to_int_uzs(value: Decimal) -> int:
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def money_fmt_uzs(value: Decimal | int | float) -> str:
    dec = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    s = f"{int(dec):,}".replace(",", " ")
    return f"{s} so‘m"


def money_fmt_usd(value: Decimal | int | float) -> str:
    dec = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"${dec:,}".replace(",", " ")


def parse_dt(s: str) -> datetime:
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")


# ------------------------------
# FSM states
# ------------------------------
class RateStates(StatesGroup):
    waiting_manual_rate = State()


class DeleteStates(StatesGroup):
    waiting_batch_id = State()


# ------------------------------
# Data classes
# ------------------------------
@dataclass
class MoneyToken:
    sign: int
    raw: str
    amount: Decimal
    currency: str  # UZS / USD
    amount_uzs: Decimal
    start: int
    end: int


@dataclass
class ParsedLine:
    author: str
    tx_at: datetime
    raw_text: str
    source_line: str
    clean_text: str
    note: str
    category: str
    counterparty: str
    income_uzs: Decimal
    expense_uzs: Decimal
    usd_total: Decimal
    uzs_total: Decimal
    tokens: List[MoneyToken]


# ------------------------------
# Keyboards
# ------------------------------
def main_keyboard() -> ReplyKeyboardMarkup:
    kb = ReplyKeyboardMarkup(resize_keyboard=True)
    kb.row(KeyboardButton("📊 Bugun"), KeyboardButton("📅 Haftalik"), KeyboardButton("🗓 Oylik"))
    kb.row(KeyboardButton("💱 Kursni belgilash"), KeyboardButton("🔄 Yangilash"))
    kb.row(KeyboardButton("📝 Text hisobot"), KeyboardButton("📤 Export"))
    kb.row(KeyboardButton("↩️ Oxirgi amalni bekor qilish"), KeyboardButton("📚 Arxiv"))
    kb.row(KeyboardButton("ℹ️ Yordam"))
    return kb


def preview_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("✅ Saqlash", callback_data="save_pending"),
        InlineKeyboardButton("❌ Bekor qilish", callback_data="cancel_pending"),
    )
    return kb


def rate_menu_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("🌐 API orqali olish", callback_data="rate_api_fetch"),
        InlineKeyboardButton("⌨️ Qo‘lda kiritish", callback_data="rate_manual_start"),
    )
    kb.add(InlineKeyboardButton("❌ Bekor qilish", callback_data="rate_cancel"))
    return kb


def rate_confirm_kb(rate_value: str, source: str) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("✅ Saqlash", callback_data=f"rate_save:{rate_value}:{source}"),
        InlineKeyboardButton("❌ Bekor qilish", callback_data="rate_cancel"),
    )
    return kb


def reset_confirm_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=2)
    kb.add(
        InlineKeyboardButton("✅ Ha, yangilash", callback_data="reset_today_confirm"),
        InlineKeyboardButton("❌ Bekor", callback_data="reset_today_cancel"),
    )
    return kb


def archive_kb() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup(row_width=3)
    kb.add(
        InlineKeyboardButton("Bugungi resetlar", callback_data="archive:today"),
        InlineKeyboardButton("So‘nggi 7 kun", callback_data="archive:week"),
        InlineKeyboardButton("So‘nggi 30 kun", callback_data="archive:month"),
    )
    return kb


# ------------------------------
# DB helpers
# ------------------------------
@asynccontextmanager
async def get_db():
    db = await aiosqlite.connect(DB_PATH)
    db.row_factory = aiosqlite.Row
    try:
        yield db
    finally:
        await db.close()


async def ensure_column(db: aiosqlite.Connection, table: str, column: str, decl: str) -> None:
    cur = await db.execute(f"PRAGMA table_info({table})")
    cols = {row[1] for row in await cur.fetchall()}
    await cur.close()
    if column not in cols:
        await db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {decl}")


async def init_db() -> None:
    async with get_db() as db:
        await db.executescript(
            """
            PRAGMA journal_mode=WAL;
            PRAGMA foreign_keys=ON;

            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                source_text TEXT NOT NULL DEFAULT '',
                summary_text TEXT NOT NULL DEFAULT '',
                item_count INTEGER NOT NULL DEFAULT 0,
                income_total_uzs INTEGER NOT NULL DEFAULT 0,
                expense_total_uzs INTEGER NOT NULL DEFAULT 0,
                net_total_uzs INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT '',
                saved_at TEXT NOT NULL DEFAULT '',
                undone_at TEXT
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER,
                user_id INTEGER NOT NULL,
                author_name TEXT NOT NULL DEFAULT '',
                tx_at TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT '',
                period_type TEXT NOT NULL DEFAULT 'live',
                period_date TEXT NOT NULL DEFAULT '',
                period_anchor TEXT NOT NULL DEFAULT '',
                is_income INTEGER NOT NULL DEFAULT 0,
                currency TEXT NOT NULL DEFAULT 'UZS',
                amount_original TEXT NOT NULL DEFAULT '0',
                amount_uzs INTEGER NOT NULL DEFAULT 0,
                income_uzs INTEGER NOT NULL DEFAULT 0,
                expense_uzs INTEGER NOT NULL DEFAULT 0,
                usd_rate TEXT NOT NULL DEFAULT '0',
                usd_total TEXT NOT NULL DEFAULT '0',
                uzs_total TEXT NOT NULL DEFAULT '0',
                category TEXT NOT NULL DEFAULT 'boshqa',
                counterparty TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT '',
                clean_text TEXT NOT NULL DEFAULT '',
                source_line TEXT NOT NULL DEFAULT '',
                raw_text TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'active',
                undone_at TEXT,
                FOREIGN KEY(batch_id) REFERENCES batches(id)
            );

            CREATE TABLE IF NOT EXISTS reset_points (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                reset_date TEXT NOT NULL,
                reset_at TEXT NOT NULL,
                note TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT ''
            );

            CREATE INDEX IF NOT EXISTS idx_records_user_tx_at ON records(user_id, tx_at);
            CREATE INDEX IF NOT EXISTS idx_records_period ON records(user_id, period_type, period_anchor);
            CREATE INDEX IF NOT EXISTS idx_records_status ON records(status);
            CREATE INDEX IF NOT EXISTS idx_batches_user_saved_at ON batches(user_id, saved_at);
            CREATE INDEX IF NOT EXISTS idx_reset_points_user_date ON reset_points(user_id, reset_date, reset_at);
            """
        )

        # forward migrations
        for column, decl in [
            ("source_text", "TEXT NOT NULL DEFAULT ''"),
            ("summary_text", "TEXT NOT NULL DEFAULT ''"),
            ("item_count", "INTEGER NOT NULL DEFAULT 0"),
            ("income_total_uzs", "INTEGER NOT NULL DEFAULT 0"),
            ("expense_total_uzs", "INTEGER NOT NULL DEFAULT 0"),
            ("net_total_uzs", "INTEGER NOT NULL DEFAULT 0"),
            ("created_at", "TEXT NOT NULL DEFAULT ''"),
            ("saved_at", "TEXT NOT NULL DEFAULT ''"),
            ("undone_at", "TEXT"),
        ]:
            await ensure_column(db, "batches", column, decl)

        for column, decl in [
            ("created_at", "TEXT NOT NULL DEFAULT ''"),
            ("period_type", "TEXT NOT NULL DEFAULT 'live'"),
            ("period_date", "TEXT NOT NULL DEFAULT ''"),
            ("period_anchor", "TEXT NOT NULL DEFAULT ''"),
            ("currency", "TEXT NOT NULL DEFAULT 'UZS'"),
            ("amount_original", "TEXT NOT NULL DEFAULT '0'"),
            ("income_uzs", "INTEGER NOT NULL DEFAULT 0"),
            ("expense_uzs", "INTEGER NOT NULL DEFAULT 0"),
            ("usd_rate", "TEXT NOT NULL DEFAULT '0'"),
            ("usd_total", "TEXT NOT NULL DEFAULT '0'"),
            ("uzs_total", "TEXT NOT NULL DEFAULT '0'"),
            ("clean_text", "TEXT NOT NULL DEFAULT ''"),
            ("source_line", "TEXT NOT NULL DEFAULT ''"),
            ("raw_text", "TEXT NOT NULL DEFAULT ''"),
            ("status", "TEXT NOT NULL DEFAULT 'active'"),
            ("undone_at", "TEXT"),
        ]:
            await ensure_column(db, "records", column, decl)

        await set_default_setting(db, "usd_rate", str(DEFAULT_USD_RATE))
        await db.commit()


async def set_default_setting(db: aiosqlite.Connection, key: str, value: str) -> None:
    cur = await db.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = await cur.fetchone()
    await cur.close()
    if row is None:
        await db.execute("INSERT INTO settings(key, value) VALUES(?, ?)", (key, value))


async def get_setting(key: str, default: str = "") -> str:
    async with get_db() as db:
        cur = await db.execute("SELECT value FROM settings WHERE key=?", (key,))
        row = await cur.fetchone()
        await cur.close()
        return row["value"] if row else default


async def set_setting_db(db: aiosqlite.Connection, key: str, value: str) -> None:
    await db.execute(
        "INSERT INTO settings(key, value) VALUES(?, ?) "
        "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
        (key, value),
    )


async def get_usd_rate() -> Decimal:
    value = await get_setting("usd_rate", str(DEFAULT_USD_RATE))
    try:
        return Decimal(value)
    except Exception:
        return DEFAULT_USD_RATE


async def set_usd_rate_db(value: Decimal, source: str = "manual") -> None:
    async with get_db() as db:
        await set_setting_db(db, "usd_rate", str(value))
        await set_setting_db(db, "usd_rate_source", source)
        await set_setting_db(db, "usd_rate_updated_at", dt_to_str(now_tz()))
        await db.commit()


async def get_usd_rate_meta() -> Tuple[str, str, str]:
    async with get_db() as db:
        cur = await db.execute("SELECT value FROM settings WHERE key='usd_rate'")
        row = await cur.fetchone()
        rate = row["value"] if row else str(DEFAULT_USD_RATE)
        await cur.close()

        cur = await db.execute("SELECT value FROM settings WHERE key='usd_rate_source'")
        row = await cur.fetchone()
        source = row["value"] if row else "default"
        await cur.close()

        cur = await db.execute("SELECT value FROM settings WHERE key='usd_rate_updated_at'")
        row = await cur.fetchone()
        updated_at = row["value"] if row else "-"
        await cur.close()

    return rate, source, updated_at


async def is_admin(user_id: int) -> bool:
    return not ADMIN_IDS or user_id in ADMIN_IDS


async def get_today_reset_anchor(user_id: int, target_date: str) -> Optional[str]:
    async with get_db() as db:
        cur = await db.execute(
            "SELECT reset_at FROM reset_points WHERE user_id=? AND reset_date=? ORDER BY reset_at DESC LIMIT 1",
            (user_id, target_date),
        )
        row = await cur.fetchone()
        await cur.close()
        return row["reset_at"] if row else None


async def create_reset_point(user_id: int, note: str) -> str:
    now = now_tz()
    reset_at = now.isoformat(sep=" ", timespec="seconds")
    async with get_db() as db:
        await db.execute(
            "INSERT INTO reset_points(user_id, reset_date, reset_at, note, created_at) VALUES(?,?,?,?,?)",
            (user_id, now.strftime("%Y-%m-%d"), reset_at, note, reset_at),
        )
        await db.commit()
    return reset_at


# ------------------------------
# Parsing logic
# ------------------------------
LINE_RE = re.compile(
    r"^\[(?P<dt>\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})\]\s*(?P<author>[^:]+):\s*(?P<text>.+?)\s*$"
)

# Number with optional leading sign, spaces/commas/dots, optional mln/ming/k, optional currency
AMOUNT_RE = re.compile(
    r"(?P<sign>[+-]?)\s*(?P<number>\d[\d\s,\.]{0,20}\d|\d)\s*(?P<mult>mln|million|ming|k)?\s*(?P<currency>\$|usd|dollar|sum|so\'m|so‘m|som)?",
    re.IGNORECASE,
)

STOPWORDS = {
    "berdim", "berildi", "berari", "olindi", "opqoldim", "olib", "qoldim", "qoyilgan",
    "qoyilmagan", "predoplata", "avans", "xizmat", "dostavka", "dokumentiga", "labo",
    "aka", "sum", "so'm", "so‘m", "usd", "dollar", "mln", "ming", "k", "va", "ham",
}

CATEGORY_KEYWORDS = {
    "dostavka": "logistika",
    "yetkazib": "logistika",
    "metan": "yoqilg‘i",
    "benzin": "yoqilg‘i",
    "gaz": "yoqilg‘i",
    "dokument": "hujjat",
    "hujjat": "hujjat",
    "avans": "avans",
    "predoplata": "avans",
    "xizmat": "xizmat",
    "resor": "ta'mir",
    "nikel": "qo‘shimcha ish",
    "temir": "material",
    "labo": "transport",
    "laboга": "transport",
    "oylik": "ish haqi",
    "maosh": "ish haqi",
}


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_amount_number(raw: str) -> Decimal:
    raw = raw.strip().replace(" ", "")
    if not raw:
        raise InvalidOperation("empty number")

    if "," in raw and "." in raw:
        # choose the last separator as decimal separator if precision <=2
        last_comma = raw.rfind(",")
        last_dot = raw.rfind(".")
        dec_pos = max(last_comma, last_dot)
        whole = re.sub(r"[\.,]", "", raw[:dec_pos])
        frac = raw[dec_pos + 1 :]
        if len(frac) <= 2:
            raw = f"{whole}.{frac}"
        else:
            raw = re.sub(r"[\.,]", "", raw)
    elif raw.count(",") >= 1 and "." not in raw:
        parts = raw.split(",")
        if len(parts[-1]) <= 2 and len(parts) == 2:
            raw = raw.replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif raw.count(".") >= 1 and "," not in raw:
        parts = raw.split(".")
        if len(parts[-1]) > 2:
            raw = raw.replace(".", "")
    return Decimal(raw)


def clean_note_text(text: str) -> str:
    text = AMOUNT_RE.sub(" ", text)
    text = re.sub(r"\b\d{2}\.\d{2}\.\d{4}\b", " ", text)
    text = re.sub(r"\b\d{2}:\d{2}\b", " ", text)
    text = normalize_spaces(text)
    return text


def guess_category(note: str, clean_text: str) -> str:
    combined = f"{note} {clean_text}".lower()
    for k, v in CATEGORY_KEYWORDS.items():
        if k in combined:
            return v
    return "boshqa"


def guess_counterparty(note: str, clean_text: str) -> str:
    tokens = re.findall(r"[A-Za-zА-Яа-яЁёЎўҚқҒғҲҳ]+", f"{note} {clean_text}")
    filtered = []
    for token in tokens:
        low = token.lower()
        if low in STOPWORDS or len(token) < 3:
            continue
        filtered.append(token)
    # keep up to 3 first meaningful tokens
    return " ".join(filtered[:3])


async def maybe_enhance_note_with_groq(clean_text: str, raw_text: str) -> Dict[str, str]:
    """Optional soft enhancement. Never blocks parsing logic."""
    if not GROQ_API_KEY:
        return {"note": clean_text[:250], "category": guess_category(clean_text, clean_text), "counterparty": guess_counterparty(clean_text, clean_text)}

    try:
        prompt = (
            "Quyidagi moliyaviy yozuv uchun qisqa, tushunarli izoh, kategoriya va kontragent ajrat. "
            "Faqat JSON qaytar. Kalitlar: note, category, counterparty. "
            "Izoh 12 ta so'zdan oshmasin. Qo'shimcha matn yozma.\n\n"
            f"RAW: {raw_text}\n"
            f"CLEAN: {clean_text}"
        )
        payload = {
            "model": GROQ_MODEL,
            "messages": [
                {"role": "system", "content": "Sen moliyaviy yozuvlarni tuzilmaga keltiradigan yordamchisan."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
        }
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json",
        }
        timeout = aiohttp.ClientTimeout(total=15)
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers=headers,
                json=payload,
            ) as resp:
                if resp.status >= 400:
                    raise RuntimeError(f"Groq status {resp.status}")
                data = await resp.json()
        content = data["choices"][0]["message"]["content"]
        obj = json.loads(content)
        note = normalize_spaces(str(obj.get("note", "")))[:250] or clean_text[:250]
        category = normalize_spaces(str(obj.get("category", "")))[:50] or guess_category(note, clean_text)
        counterparty = normalize_spaces(str(obj.get("counterparty", "")))[:80] or guess_counterparty(note, clean_text)
        return {"note": note, "category": category, "counterparty": counterparty}
    except Exception:
        logger.warning("Groq enhancement failed", exc_info=True)
        return {
            "note": clean_text[:250],
            "category": guess_category(clean_text, clean_text),
            "counterparty": guess_counterparty(clean_text, clean_text),
        }


async def parse_finance_line(line: str, usd_rate: Decimal) -> Optional[ParsedLine]:
    m = LINE_RE.match(line.strip())
    if not m:
        return None

    dt = datetime.strptime(m.group("dt"), "%d.%m.%Y %H:%M").replace(tzinfo=TZ)
    author = normalize_spaces(m.group("author"))
    raw_text = normalize_spaces(m.group("text"))

    tokens: List[MoneyToken] = []
    for match in AMOUNT_RE.finditer(raw_text):
        raw = match.group(0)
        number = match.group("number")
        sign_s = match.group("sign") or ""
        mult = (match.group("mult") or "").lower()
        currency_s = (match.group("currency") or "").lower()

        # skip clear non-money garbage like 018, dates fragments, too short leading zero numbers without unit/currency/sign
        if not sign_s and not mult and not currency_s and re.fullmatch(r"0\d{1,3}", number.replace(" ", "")):
            continue

        try:
            amount = parse_amount_number(number)
        except Exception:
            continue

        if amount <= 0:
            continue

        if mult in {"mln", "million"}:
            amount *= Decimal("1000000")
        elif mult in {"ming", "k"}:
            amount *= Decimal("1000")

        currency = "USD" if currency_s in {"$", "usd", "dollar"} else "UZS"
        sign = 1 if sign_s == "+" else -1
        amount_uzs = amount * usd_rate if currency == "USD" else amount

        tokens.append(
            MoneyToken(
                sign=sign,
                raw=raw,
                amount=amount,
                currency=currency,
                amount_uzs=amount_uzs,
                start=match.start(),
                end=match.end(),
            )
        )

    if not tokens:
        return None

    income = Decimal("0")
    expense = Decimal("0")
    usd_total = Decimal("0")
    uzs_total = Decimal("0")
    for token in tokens:
        if token.currency == "USD":
            usd_total += token.amount
        else:
            uzs_total += token.amount
        if token.sign > 0:
            income += token.amount_uzs
        else:
            expense += token.amount_uzs

    clean_text = clean_note_text(raw_text)
    enhanced = await maybe_enhance_note_with_groq(clean_text, raw_text)
    note = enhanced["note"] or clean_text or raw_text[:250]
    category = enhanced["category"] or guess_category(note, clean_text)
    counterparty = enhanced["counterparty"] or guess_counterparty(note, clean_text)

    return ParsedLine(
        author=author,
        tx_at=dt,
        raw_text=raw_text,
        source_line=line.strip(),
        clean_text=clean_text,
        note=note[:250],
        category=category[:50] or "boshqa",
        counterparty=counterparty[:80],
        income_uzs=income,
        expense_uzs=expense,
        usd_total=usd_total,
        uzs_total=uzs_total,
        tokens=tokens,
    )


async def parse_text_blob(text: str, usd_rate: Decimal) -> List[ParsedLine]:
    lines = [normalize_spaces(x) for x in text.splitlines() if normalize_spaces(x)]
    parsed: List[ParsedLine] = []
    for line in lines:
        item = await parse_finance_line(line, usd_rate)
        if item:
            parsed.append(item)
    return parsed


# ------------------------------
# Reporting & persistence
# ------------------------------
def summarize_items(items: Sequence[ParsedLine]) -> Dict[str, Decimal]:
    income = sum((x.income_uzs for x in items), Decimal("0"))
    expense = sum((x.expense_uzs for x in items), Decimal("0"))
    return {
        "income": income,
        "expense": expense,
        "net": income - expense,
    }


async def save_pending_batch(pending: Dict[str, Any]) -> int:
    user_id = int(pending["user_id"])
    items: List[ParsedLine] = pending["items"]
    source_text = pending["source_text"]
    created_at = now_tz().isoformat(sep=" ", timespec="seconds")
    stats = summarize_items(items)
    today_date = now_tz().strftime("%Y-%m-%d")
    reset_anchor = await get_today_reset_anchor(user_id, today_date) or ""
    usd_rate = await get_usd_rate()

    async with get_db() as db:
        cur = await db.execute(
            """
            INSERT INTO batches(
                user_id, source_text, summary_text, item_count,
                income_total_uzs, expense_total_uzs, net_total_uzs,
                created_at, saved_at
            ) VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (
                user_id,
                source_text,
                pending["summary_text"],
                len(items),
                decimal_to_int_uzs(stats["income"]),
                decimal_to_int_uzs(stats["expense"]),
                decimal_to_int_uzs(stats["net"]),
                created_at,
                created_at,
            ),
        )
        batch_id = cur.lastrowid
        await cur.close()

        for item in items:
            is_income = 1 if item.income_uzs > 0 else 0
            amount_uzs = item.income_uzs if item.income_uzs > 0 else item.expense_uzs
            amount_original = str(item.usd_total if item.usd_total > 0 else item.uzs_total)
            currency = "USD" if item.usd_total > 0 and item.uzs_total == 0 else "UZS"
            period_anchor = item.tx_at.strftime("%Y-%m-%d")
            await db.execute(
                """
                INSERT INTO records(
                    batch_id, user_id, author_name, tx_at, created_at,
                    period_type, period_date, period_anchor,
                    is_income, currency, amount_original, amount_uzs,
                    income_uzs, expense_uzs, usd_rate, usd_total, uzs_total,
                    category, counterparty, note, clean_text, source_line, raw_text,
                    status
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?, 'active')
                """,
                (
                    batch_id,
                    user_id,
                    item.author,
                    item.tx_at.isoformat(sep=" ", timespec="seconds"),
                    created_at,
                    "live",
                    today_date,
                    reset_anchor,
                    is_income,
                    currency,
                    amount_original,
                    decimal_to_int_uzs(amount_uzs),
                    decimal_to_int_uzs(item.income_uzs),
                    decimal_to_int_uzs(item.expense_uzs),
                    str(usd_rate),
                    str(item.usd_total),
                    str(item.uzs_total),
                    item.category,
                    item.counterparty,
                    item.note,
                    item.clean_text,
                    item.source_line,
                    item.raw_text,
                ),
            )
        await db.commit()
    return int(batch_id)


async def undo_last_batch(user_id: int) -> Optional[int]:
    now_s = now_tz().isoformat(sep=" ", timespec="seconds")
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM batches WHERE user_id=? AND undone_at IS NULL ORDER BY id DESC LIMIT 1",
            (user_id,),
        )
        row = await cur.fetchone()
        await cur.close()
        if not row:
            return None
        batch_id = int(row["id"])
        await db.execute("UPDATE batches SET undone_at=? WHERE id=?", (now_s, batch_id))
        await db.execute(
            "UPDATE records SET status='undone', undone_at=? WHERE batch_id=? AND status='active'",
            (now_s, batch_id),
        )
        await db.commit()
        return batch_id


async def fetch_records(
    user_id: int,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    today_live_only: bool = False,
    limit: Optional[int] = None,
) -> List[aiosqlite.Row]:
    sql = "SELECT * FROM records WHERE user_id=? AND status='active'"
    params: List[Any] = [user_id]

    if date_from is not None:
        sql += " AND tx_at >= ?"
        params.append(date_from.isoformat(sep=" ", timespec="seconds"))
    if date_to is not None:
        sql += " AND tx_at < ?"
        params.append(date_to.isoformat(sep=" ", timespec="seconds"))
    if today_live_only:
        anchor = await get_today_reset_anchor(user_id, now_tz().strftime("%Y-%m-%d")) or ""
        sql += " AND period_date=? AND period_anchor=?"
        params.extend([now_tz().strftime("%Y-%m-%d"), anchor])
    sql += " ORDER BY tx_at DESC, id DESC"
    if limit:
        sql += f" LIMIT {int(limit)}"

    async with get_db() as db:
        cur = await db.execute(sql, tuple(params))
        rows = await cur.fetchall()
        await cur.close()
        return rows


async def compute_summary(user_id: int, mode: str) -> Dict[str, Any]:
    now = now_tz()
    if mode == "today":
        rows = await fetch_records(user_id, today_live_only=True)
        title = "Bugungi hisobot"
    elif mode == "week":
        start = datetime(now.year, now.month, now.day, tzinfo=TZ) - timedelta(days=6)
        end = datetime(now.year, now.month, now.day, tzinfo=TZ) + timedelta(days=1)
        rows = await fetch_records(user_id, date_from=start, date_to=end)
        title = "Haftalik hisobot"
    elif mode == "month":
        start = datetime(now.year, now.month, 1, tzinfo=TZ)
        if now.month == 12:
            end = datetime(now.year + 1, 1, 1, tzinfo=TZ)
        else:
            end = datetime(now.year, now.month + 1, 1, tzinfo=TZ)
        rows = await fetch_records(user_id, date_from=start, date_to=end)
        title = "Oylik hisobot"
    else:
        rows = await fetch_records(user_id)
        title = "Umumiy hisobot"

    income = sum(Decimal(row["income_uzs"]) for row in rows)
    expense = sum(Decimal(row["expense_uzs"]) for row in rows)
    balance = income - expense

    by_cat: Dict[str, Decimal] = {}
    for row in rows:
        category = row["category"] or "boshqa"
        by_cat[category] = by_cat.get(category, Decimal("0")) + Decimal(row["amount_uzs"])

    return {
        "title": title,
        "rows": rows,
        "income": income,
        "expense": expense,
        "balance": balance,
        "by_cat": sorted(by_cat.items(), key=lambda x: x[1], reverse=True),
    }


def build_summary_text(data: Dict[str, Any]) -> str:
    rows = data["rows"]
    parts = [
        f"<b>{data['title']}</b>",
        f"Yozuvlar soni: <b>{len(rows)}</b>",
        f"Kirim: <b>{money_fmt_uzs(data['income'])}</b>",
        f"Chiqim: <b>{money_fmt_uzs(data['expense'])}</b>",
        f"Qoldiq: <b>{money_fmt_uzs(data['balance'])}</b>",
    ]
    if data["by_cat"]:
        parts.append("\n<b>Kategoriyalar:</b>")
        for category, amount in data["by_cat"][:7]:
            parts.append(f"• {category}: {money_fmt_uzs(amount)}")
    if rows:
        parts.append("\n<b>So‘nggi yozuvlar:</b>")
        for row in rows[:5]:
            sign = "+" if row["is_income"] else "-"
            parts.append(
                f"• {dt_to_str(parse_dt(row['tx_at']))} | {sign}{money_fmt_uzs(row['amount_uzs'])} | {row['note']}"
            )
    return "\n".join(parts)


def build_records_text(rows: Sequence[aiosqlite.Row], title: str = "Yozuvlar") -> str:
    parts = [f"<b>{title}</b>"]
    if not rows:
        parts.append("Yozuv topilmadi.")
        return "\n".join(parts)
    for idx, row in enumerate(rows, start=1):
        sign = "+" if row["is_income"] else "-"
        parts.append(
            f"\n<b>{idx}.</b> #{row['id']} | {dt_to_str(parse_dt(row['tx_at']))}\n"
            f"{sign}{money_fmt_uzs(row['amount_uzs'])}\n"
            f"Izoh: {row['note']}\n"
            f"Kategoriya: {row['category']}\n"
            f"Kontragent: {row['counterparty'] or '-'}"
        )
    return "\n".join(parts)


def _autosize_ws(ws) -> None:
    for col_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)


def build_excel_bytes(rows: Sequence[aiosqlite.Row], summary: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    dashboard_rows = [
        ("Hisobot", summary["title"]),
        ("Yozuvlar soni", len(rows)),
        ("Kirim", decimal_to_int_uzs(summary["income"])),
        ("Chiqim", decimal_to_int_uzs(summary["expense"])),
        ("Qoldiq", decimal_to_int_uzs(summary["balance"])),
        ("Yaratilgan vaqt", dt_to_str(now_tz())),
    ]
    for r_idx, (k, v) in enumerate(dashboard_rows, start=1):
        ws_dash.cell(r_idx, 1, k)
        ws_dash.cell(r_idx, 2, v)
    for c in ws_dash[1]:
        c.fill = header_fill
        c.font = header_font
    _autosize_ws(ws_dash)

    def fill_sheet(name: str, subset: Sequence[aiosqlite.Row]) -> None:
        ws = wb.create_sheet(name)
        headers = [
            "ID",
            "Sana va vaqt",
            "Tur",
            "Summa (UZS)",
            "Valyuta",
            "Original summa",
            "USD kurs",
            "Kategoriya",
            "Kontragent",
            "Izoh",
            "Tozalangan matn",
            "Muallif",
            "Asl satr",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in subset:
            ws.append([
                row["id"],
                dt_to_str(parse_dt(row["tx_at"])),
                "Kirim" if row["is_income"] else "Chiqim",
                int(row["amount_uzs"]),
                row["currency"],
                row["amount_original"],
                row["usd_rate"],
                row["category"],
                row["counterparty"],
                row["note"],
                row["clean_text"],
                row["author_name"],
                row["source_line"],
            ])
        _autosize_ws(ws)

    fill_sheet("Barcha", rows)
    fill_sheet("Kirim", [r for r in rows if r["is_income"]])
    fill_sheet("Chiqim", [r for r in rows if not r["is_income"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv_bytes(rows: Sequence[aiosqlite.Row]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "id", "tx_at", "type", "amount_uzs", "currency", "amount_original", "usd_rate",
        "category", "counterparty", "note", "clean_text", "author_name", "source_line"
    ])
    for row in rows:
        writer.writerow([
            row["id"], dt_to_str(parse_dt(row["tx_at"])), "income" if row["is_income"] else "expense",
            row["amount_uzs"], row["currency"], row["amount_original"], row["usd_rate"],
            row["category"], row["counterparty"], row["note"], row["clean_text"],
            row["author_name"], row["source_line"],
        ])
    return buf.getvalue().encode("utf-8-sig")


def build_plaintext_report(rows: Sequence[aiosqlite.Row], summary: Dict[str, Any]) -> bytes:
    lines = [build_summary_text(summary), "", build_records_text(rows, "To‘liq yozuvlar")]
    return "\n".join(lines).encode("utf-8")


# ------------------------------
# Rate API
# ------------------------------
async def fetch_cbu_usd_rate() -> Dict[str, Any]:
    timeout = aiohttp.ClientTimeout(total=20)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(RATE_API_URL) as resp:
            resp.raise_for_status()
            data = await resp.json()
    if not isinstance(data, list) or not data:
        raise ValueError("API bo‘sh javob qaytardi")
    item = data[0]
    raw_rate = str(item.get("Rate", "")).strip().replace(",", ".")
    if not raw_rate:
        raise ValueError("Rate topilmadi")
    return {
        "rate": Decimal(raw_rate),
        "date": str(item.get("Date", "")).strip(),
        "ccy": item.get("Ccy", "USD"),
        "name": item.get("CcyNm_UZ", "AQSH dollari"),
    }


# ------------------------------
# Message rendering
# ------------------------------
def build_preview_text(items: Sequence[ParsedLine], usd_rate: Decimal) -> str:
    stats = summarize_items(items)
    parts = [
        "<b>Saqlashdan oldin preview</b>",
        f"Topilgan yozuvlar: <b>{len(items)}</b>",
        f"USD kurs: <b>{money_fmt_uzs(usd_rate)}</b>",
        f"Kirim: <b>{money_fmt_uzs(stats['income'])}</b>",
        f"Chiqim: <b>{money_fmt_uzs(stats['expense'])}</b>",
        f"Qoldiq: <b>{money_fmt_uzs(stats['net'])}</b>",
        "\n<b>Yozuvlar:</b>",
    ]
    for idx, item in enumerate(items[:10], start=1):
        sign = "+" if item.income_uzs > 0 else "-"
        amount = item.income_uzs if item.income_uzs > 0 else item.expense_uzs
        parts.append(
            f"{idx}. {dt_to_str(item.tx_at)} | {sign}{money_fmt_uzs(amount)} | {item.note} | {item.category}"
        )
    if len(items) > 10:
        parts.append(f"... yana {len(items) - 10} ta yozuv")
    return "\n".join(parts)


# ------------------------------
# Handlers
# ------------------------------
@dp.errors_handler()
async def global_error_handler(update, exception):
    logger.error("Unhandled error: %s\n%s", exception, traceback.format_exc())
    return True


@dp.message_handler(commands=["start", "menu"])
async def cmd_start(message: types.Message):
    if not await is_admin(message.from_user.id):
        return await message.answer("Ruxsat yo‘q.")
    await message.answer(
        "<b>Xisobot Bot Pro</b>\n\n"
        "Yuborish formati:\n"
        "[27.03.2026 12:01] Алишеров Орифжон: 250$+517 ming azam aka labo berari olindi\n\n"
        "Qoidalar:\n"
        "• <b>+</b> bilan boshlangan summa — <b>kirim</b>\n"
        "• <b>+</b> bo‘lmasa — <b>chiqim</b>\n"
        "• <b>mln</b> = million\n"
        "• <b>ming</b> yoki <b>k</b> = ming\n"
        "• <b>$</b> yoki <b>usd</b> — dollar, kurs bo‘yicha so‘mga o‘tkaziladi\n\n"
        "Bot avval preview ko‘rsatadi, keyin saqlashni so‘raydi.",
        reply_markup=main_keyboard(),
    )


@dp.message_handler(commands=["help"])
@dp.message_handler(lambda m: m.text == "ℹ️ Yordam")
async def cmd_help(message: types.Message):
    await message.answer(
        "<b>Asosiy buyruqlar</b>\n"
        "/rate - joriy kurs\n"
        "/rate 12750 - kursni qo‘lda o‘rnatish\n"
        "/stats - umumiy hisobot\n"
        "/records 10 - oxirgi 10 yozuv\n"
        "/undo - oxirgi batchni bekor qilish\n"
        "/export - hisobot fayllari\n\n"
        "Tugmalar orqali ham hammasi ishlaydi.",
        reply_markup=main_keyboard(),
    )


@dp.message_handler(commands=["rate"])
async def cmd_rate(message: types.Message):
    if not await is_admin(message.from_user.id):
        return
    args = (message.get_args() or "").strip()
    if not args:
        rate, source, updated_at = await get_usd_rate_meta()
        return await message.answer(
            f"💱 Joriy USD kursi: <b>{rate}</b>\n"
            f"Manba: <b>{source}</b>\n"
            f"Yangilangan: <b>{updated_at}</b>\n\n"
            f"Qo‘lda yangilash: <code>/rate 12750</code>",
            reply_markup=main_keyboard(),
        )
    try:
        value = Decimal(args.replace(",", "").replace(" ", ""))
        if value <= 0:
            raise ValueError
    except Exception:
        return await message.answer("❌ Noto‘g‘ri format. Misol: /rate 12750")
    await set_usd_rate_db(value, source="manual")
    await message.answer(f"✅ Yangi USD kurs saqlandi: <b>{value}</b>", reply_markup=main_keyboard())


@dp.message_handler(lambda m: m.text == "💱 Kursni belgilash")
async def rate_menu_handler(message: types.Message):
    if not await is_admin(message.from_user.id):
        return
    rate, source, updated_at = await get_usd_rate_meta()
    await message.answer(
        f"💱 Joriy USD kursi: <b>{rate}</b>\n"
        f"Manba: <b>{source}</b>\n"
        f"Yangilangan: <b>{updated_at}</b>\n\n"
        "Tanlang:\n"
        "• API orqali avtomatik olish\n"
        "• Qo‘lda kiritish\n\n"
        "Qo‘lda format: <code>12750</code> yoki <code>12,750</code>",
        reply_markup=rate_menu_kb(),
    )


@dp.callback_query_handler(lambda c: c.data == "rate_api_fetch")
async def rate_api_fetch_callback(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("Ruxsat yo‘q", show_alert=True)
    try:
        info = await fetch_cbu_usd_rate()
        rate_str = str(info["rate"])
        await call.message.edit_text(
            f"🌐 API natijasi\n"
            f"Valyuta: {info['name']} ({info['ccy']})\n"
            f"Kurs: <b>{rate_str}</b>\n"
            f"Sana: {info['date']}\n\n"
            f"Shu kursni saqlaysizmi?",
            reply_markup=rate_confirm_kb(rate_str, "api"),
        )
        await call.answer()
    except Exception as e:
        await call.answer("API xatosi", show_alert=True)
        await call.message.answer(f"❌ API dan kurs olinmadi: {e}")


@dp.callback_query_handler(lambda c: c.data == "rate_manual_start")
async def rate_manual_start_callback(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("Ruxsat yo‘q", show_alert=True)
    await RateStates.waiting_manual_rate.set()
    await call.answer()
    await call.message.answer(
        "⌨️ Yangi USD kursini yuboring.\n\n"
        "Misollar:\n"
        "• 12750\n"
        "• 12,750\n"
        "• 12750.50"
    )


@dp.message_handler(state=RateStates.waiting_manual_rate, content_types=ContentType.TEXT)
async def rate_manual_input_handler(message: types.Message, state: FSMContext):
    if not await is_admin(message.from_user.id):
        await state.finish()
        return
    raw = (message.text or "").strip().replace(" ", "").replace(",", "")
    try:
        value = Decimal(raw)
        if value <= 0:
            raise ValueError
    except Exception:
        return await message.answer(
            "❌ Noto‘g‘ri format. To‘g‘ri misollar:\n• 12750\n• 12,750\n• 12750.50"
        )
    await state.finish()
    await message.answer(
        f"💱 Kiritilgan kurs: <b>{value}</b>\n\nShuni saqlaysizmi?",
        reply_markup=rate_confirm_kb(str(value), "manual"),
    )


@dp.callback_query_handler(lambda c: c.data.startswith("rate_save:"), state="*")
async def rate_save_callback(call: CallbackQuery, state: FSMContext):
    await state.finish()
    if not await is_admin(call.from_user.id):
        return await call.answer("Ruxsat yo‘q", show_alert=True)
    _, rate_value, source = call.data.split(":", 2)
    try:
        value = Decimal(rate_value)
        await set_usd_rate_db(value, source=source)
        await call.message.edit_text(f"✅ USD kurs saqlandi: <b>{value}</b>\nManba: <b>{source}</b>")
        await call.answer("Saqlandi")
    except Exception as e:
        await call.answer("Xato", show_alert=True)
        await call.message.answer(f"❌ Saqlashda xato: {e}")


@dp.callback_query_handler(lambda c: c.data == "rate_cancel", state="*")
async def rate_cancel_callback(call: CallbackQuery, state: FSMContext):
    await state.finish()
    await call.answer("Bekor qilindi")
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass


@dp.message_handler(lambda m: m.text == "🔄 Yangilash")
async def refresh_today_handler(message: types.Message):
    if not await is_admin(message.from_user.id):
        return
    await message.answer(
        "Bugungi hisobni 0 dan boshlamoqchimisiz?\n"
        "Eski arxiv saqlanadi, faqat bugungi live hisob yangi anchor bilan boshlanadi.",
        reply_markup=reset_confirm_kb(),
    )


@dp.callback_query_handler(lambda c: c.data == "reset_today_confirm")
async def reset_today_confirm(call: CallbackQuery):
    if not await is_admin(call.from_user.id):
        return await call.answer("Ruxsat yo‘q", show_alert=True)
    reset_at = await create_reset_point(call.from_user.id, "manual reset")
    await call.answer("Yangilandi")
    await call.message.edit_text(
        f"✅ Bugungi hisob yangilandi.\n"
        f"Yangi boshlanish vaqti: <b>{dt_to_str(parse_dt(reset_at))}</b>\n"
        f"Eski yozuvlar arxivda saqlanadi."
    )


@dp.callback_query_handler(lambda c: c.data == "reset_today_cancel")
async def reset_today_cancel(call: CallbackQuery):
    await call.answer("Bekor qilindi")
    await call.message.edit_text("❌ Yangilash bekor qilindi.")


@dp.message_handler(lambda m: m.text == "📊 Bugun")
async def today_report_handler(message: types.Message):
    data = await compute_summary(message.from_user.id, "today")
    await message.answer(build_summary_text(data), reply_markup=main_keyboard())


@dp.message_handler(lambda m: m.text == "📅 Haftalik")
async def week_report_handler(message: types.Message):
    data = await compute_summary(message.from_user.id, "week")
    await message.answer(build_summary_text(data), reply_markup=main_keyboard())


@dp.message_handler(lambda m: m.text == "🗓 Oylik")
async def month_report_handler(message: types.Message):
    data = await compute_summary(message.from_user.id, "month")
    await message.answer(build_summary_text(data), reply_markup=main_keyboard())


@dp.message_handler(lambda m: m.text == "📝 Text hisobot")
async def text_report_handler(message: types.Message):
    data = await compute_summary(message.from_user.id, "month")
    rows = data["rows"]
    bio = io.BytesIO(build_plaintext_report(rows, data))
    bio.name = f"hisobot_{today_str()}.txt"
    await message.answer_document(InputFile(bio), caption="Text hisobot tayyor")


@dp.message_handler(commands=["stats"])
async def cmd_stats(message: types.Message):
    data = await compute_summary(message.from_user.id, "all")
    await message.answer(build_summary_text(data), reply_markup=main_keyboard())


@dp.message_handler(commands=["records"])
async def cmd_records(message: types.Message):
    args = (message.get_args() or "10").strip()
    try:
        limit = max(1, min(int(args), 50))
    except Exception:
        limit = 10
    rows = await fetch_records(message.from_user.id, limit=limit)
    await message.answer(build_records_text(rows, f"Oxirgi {limit} yozuv"), reply_markup=main_keyboard())


@dp.message_handler(commands=["undo"])
@dp.message_handler(lambda m: m.text == "↩️ Oxirgi amalni bekor qilish")
async def cmd_undo(message: types.Message):
    batch_id = await undo_last_batch(message.from_user.id)
    if batch_id is None:
        await message.answer("Bekor qilinadigan batch topilmadi.", reply_markup=main_keyboard())
    else:
        await message.answer(f"✅ Oxirgi batch bekor qilindi: <b>#{batch_id}</b>", reply_markup=main_keyboard())


@dp.message_handler(lambda m: m.text == "📚 Arxiv")
async def archive_handler(message: types.Message):
    await message.answer("Arxiv ko‘rish oynasi:", reply_markup=archive_kb())


@dp.callback_query_handler(lambda c: c.data.startswith("archive:"))
async def archive_callback(call: CallbackQuery):
    scope = call.data.split(":", 1)[1]
    async with get_db() as db:
        if scope == "today":
            cur = await db.execute(
                "SELECT * FROM reset_points WHERE user_id=? AND reset_date=? ORDER BY reset_at DESC LIMIT 10",
                (call.from_user.id, now_tz().strftime("%Y-%m-%d")),
            )
        elif scope == "week":
            cur = await db.execute(
                "SELECT * FROM reset_points WHERE user_id=? AND reset_at>=? ORDER BY reset_at DESC LIMIT 20",
                ((call.from_user.id), (now_tz() - timedelta(days=7)).isoformat(sep=' ', timespec='seconds')),
            )
        else:
            cur = await db.execute(
                "SELECT * FROM reset_points WHERE user_id=? AND reset_at>=? ORDER BY reset_at DESC LIMIT 50",
                ((call.from_user.id), (now_tz() - timedelta(days=30)).isoformat(sep=' ', timespec='seconds')),
            )
        rows = await cur.fetchall()
        await cur.close()
    if not rows:
        return await call.message.edit_text("Arxiv topilmadi.")
    text = ["<b>Reset arxivi</b>"]
    for row in rows:
        text.append(f"• {dt_to_str(parse_dt(row['reset_at']))} | {row['note'] or '-'}")
    await call.message.edit_text("\n".join(text[:60]))


@dp.message_handler(commands=["export"])
@dp.message_handler(lambda m: m.text == "📤 Export")
async def export_handler(message: types.Message):
    data = await compute_summary(message.from_user.id, "month")
    rows = data["rows"]
    if not rows:
        return await message.answer("Export uchun yozuv topilmadi.")

    xlsx = io.BytesIO(build_excel_bytes(rows, data))
    xlsx.name = f"hisobot_{today_str()}.xlsx"
    csv_bytes = io.BytesIO(build_csv_bytes(rows))
    csv_bytes.name = f"hisobot_{today_str()}.csv"
    txt_bytes = io.BytesIO(build_plaintext_report(rows, data))
    txt_bytes.name = f"hisobot_{today_str()}.txt"

    await message.answer_document(InputFile(xlsx), caption="Excel hisobot")
    await message.answer_document(InputFile(csv_bytes), caption="CSV hisobot")
    await message.answer_document(InputFile(txt_bytes), caption="Text hisobot")


@dp.message_handler(content_types=ContentType.TEXT, state=None)
async def parse_text_message(message: types.Message):
    if not await is_admin(message.from_user.id):
        return

    text = (message.text or "").strip()
    reserved = {
        "📊 Bugun", "📅 Haftalik", "🗓 Oylik", "💱 Kursni belgilash",
        "🔄 Yangilash", "📝 Text hisobot", "📤 Export", "↩️ Oxirgi amalni bekor qilish",
        "📚 Arxiv", "ℹ️ Yordam",
    }
    if text.startswith("/") or text in reserved:
        return

    usd_rate = await get_usd_rate()
    items = await parse_text_blob(text, usd_rate)
    if not items:
        return await message.answer(
            "Yaroqli yozuv topilmadi. Format misoli:\n"
            "<code>[27.03.2026 12:01] Алишеров Орифжон: 250$+517 ming azam aka labo berari olindi</code>",
            reply_markup=main_keyboard(),
        )

    summary_text = build_preview_text(items, usd_rate)
    PENDING_BATCHES[message.from_user.id] = {
        "user_id": message.from_user.id,
        "items": items,
        "source_text": text,
        "summary_text": summary_text,
    }
    await message.answer(summary_text, reply_markup=preview_kb())


@dp.callback_query_handler(lambda c: c.data == "save_pending")
async def callback_save(call: CallbackQuery):
    pending = PENDING_BATCHES.get(call.from_user.id)
    if not pending:
        return await call.answer("Saqlanadigan preview topilmadi", show_alert=True)
    batch_id = await save_pending_batch(pending)
    PENDING_BATCHES.pop(call.from_user.id, None)
    await call.answer("Saqlandi")
    await call.message.edit_text(f"✅ Batch saqlandi: <b>#{batch_id}</b>\n\n{pending['summary_text']}")


@dp.callback_query_handler(lambda c: c.data == "cancel_pending")
async def callback_cancel(call: CallbackQuery):
    PENDING_BATCHES.pop(call.from_user.id, None)
    await call.answer("Bekor qilindi")
    await call.message.edit_text("❌ Preview bekor qilindi.")


async def on_startup(_: Dispatcher) -> None:
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError("TELEGRAM_BOT_TOKEN topilmadi")
    await init_db()
    logger.info("Bot ishga tushdi")


def main() -> None:
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)


if __name__ == "__main__":
    main()
